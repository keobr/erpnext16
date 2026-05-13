"""
dn_mailer/xlsx_builder.py
Строит XLSX-файл для Delivery Note и возвращает bytes.
Не содержит frappe-зависимостей — легко тестировать отдельно.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── цвета ────────────────────────────────────────────────────
_DARK  = "1F3864"
_WHITE = "FFFFFF"
_ZEBRA = "EEF2FF"
_GREY  = "888888"
_BORDER_COLOR = "BDBDBD"


def _side():
    return Side(style="thin", color=_BORDER_COLOR)


def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Arial", size=size, bold=bold,
                color=color, italic=italic)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _fill(color):
    return PatternFill("solid", start_color=color)


def _apply(cell, font=None, fill=None, alignment=None,
           border=None, number_format=None):
    if font:            cell.font          = font
    if fill:            cell.fill          = fill
    if alignment:       cell.alignment     = alignment
    if border:          cell.border        = border
    if number_format:   cell.number_format = number_format


# ── колонки таблицы ───────────────────────────────────────────
_COLUMNS = [
    # (заголовок,       ключ,          ширина, числовой?)
    ("№",              None,           6,      True),
    ("Код товара",     "item_code",    18,     False),
    ("Наименование",   "item_name",    42,     False),
    ("Кол-во",         "qty",          10,     True),
    ("Ед. изм.",       "uom",          10,     False),
    ("Склад",          "warehouse",    20,     False),
    ("Цена",           "rate",         14,     True),
    ("Сумма",          "amount",       16,     True),
]


# ── публичная функция ─────────────────────────────────────────
def build_xlsx(doc, company_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Note"

    _set_col_widths(ws)
    _write_header(ws, doc, company_name)
    _write_meta(ws, doc)
    last_row = _write_table(ws, doc)
    _write_footer(ws, doc, last_row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── внутренние функции ────────────────────────────────────────
def _set_col_widths(ws):
    for i, (_, _, width, _) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_header(ws, doc, company_name):
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"{company_name}  |  Накладная на отгрузку"
    _apply(c, font=_font(bold=True, size=14, color=_DARK),
           alignment=_align("left"))


def _write_meta(ws, doc):
    meta = [
        ("Номер накладной:", doc.name),
        ("Клиент:",          doc.customer),
        ("Дата отгрузки:",   str(doc.posting_date)),
        ("Адрес доставки:",  getattr(doc, "shipping_address", None) or "—"),
    ]
    for idx, (label, value) in enumerate(meta, start=3):
        ws.row_dimensions[idx].height = 17
        lc = ws.cell(row=idx, column=1, value=label)
        _apply(lc, font=_font(bold=True, color="444444"),
               alignment=_align("left"))
        vc = ws.cell(row=idx, column=2, value=value)
        _apply(vc, font=_font(), alignment=_align("left"))
        ws.merge_cells(start_row=idx, start_column=2,
                       end_row=idx,   end_column=8)


def _write_table(ws, doc) -> int:
    HEADER_ROW = 8
    ws.row_dimensions[HEADER_ROW].height = 28

    for col, (title, _, _, _) in enumerate(_COLUMNS, start=1):
        c = ws.cell(row=HEADER_ROW, column=col, value=title)
        _apply(c,
               font=_font(bold=True, color=_WHITE),
               fill=_fill(_DARK),
               alignment=_align("center", wrap=True),
               border=_border())

    for i, item in enumerate(doc.items, start=1):
        r = HEADER_ROW + i
        ws.row_dimensions[r].height = 16

        qty    = item.qty    or 0
        rate   = item.rate   or 0
        amount = item.amount or (qty * rate)

        row_values = [i, item.item_code, item.item_name,
                      qty, item.uom, item.warehouse, rate, amount]

        for col, (_, _, _, is_num) in enumerate(_COLUMNS, start=1):
            c = ws.cell(row=r, column=col, value=row_values[col - 1])
            _apply(c,
                   font=_font(),
                   alignment=_align("right" if is_num else "left"),
                   border=_border(),
                   number_format="#,##0.00" if is_num and col > 3 else None)
            if i % 2 == 0:
                c.fill = _fill(_ZEBRA)

    return HEADER_ROW + len(doc.items)


def _write_footer(ws, doc, last_data_row):
    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 20

    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row,   end_column=7)
    lbl = ws.cell(row=total_row, column=1, value="ИТОГО:")
    _apply(lbl, font=_font(bold=True, color=_DARK),
           alignment=_align("right"), border=_border())

    total_amount = sum(
        (item.amount or (item.qty or 0) * (item.rate or 0))
        for item in doc.items
    )
    amt = ws.cell(row=total_row, column=8, value=total_amount)
    _apply(amt, font=_font(bold=True, color=_DARK),
           alignment=_align("right"), border=_border(),
           number_format="#,##0.00")

    sig_row = total_row + 2
    ws.merge_cells(f"A{sig_row}:H{sig_row}")
    note = ws[f"A{sig_row}"]
    total_qty = sum(item.qty or 0 for item in doc.items)
    note.value = (
        f"Документ сформирован автоматически  |  "
        f"Позиций: {len(doc.items)}  |  "
        f"Итого единиц: {total_qty}"
    )
    _apply(note, font=_font(italic=True, size=9, color=_GREY),
           alignment=_align("left"))
